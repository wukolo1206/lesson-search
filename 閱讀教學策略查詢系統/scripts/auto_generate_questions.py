"""
auto_generate_questions.py
Automatically generates transfer questions for all rows in 學習遷移題目
using the Gemini API, calibrated to the original question's difficulty.

Usage: python auto_generate_questions.py
"""

import pandas as pd
from openpyxl import load_workbook
from google import genai
import time

# ===================================
# 設定
# ===================================
API_KEY = "AIzaSyAEuisDpH_-QKKnz3EabCGxUrWDTxL-ei8"
EXCEL_FILE = r"d:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
SHEET_NAME = "學習遷移題目"
MODEL_NAME = "gemini-2.0-flash-lite"

# 每次處理幾筆後儲存一次（防止中途出錯遺失進度）
SAVE_EVERY = 10

# 設為 0 代表從頭開始，設為其他數字代表從第 N 列繼續
START_FROM_ROW = 30

# 每筆之間等幾秒（免費方案上限 15次/分鐘，6 秒 ≒ 10次/分鐘，安全範圍）
SLEEP_SECONDS = 6

# 遇到 429 限速時，最多重試幾次
MAX_RETRIES = 3

# ===================================
# 初始化 Gemini
# ===================================
client = genai.Client(api_key=API_KEY)



def get_strategy_instruction(strategy: str, process: str) -> str:
    strategy = strategy.strip()
    process = process.strip()

    vocab_keywords = ['拆字', '解詞', '詞義', '上下文', '推詞', '字義']
    if any(k in strategy for k in vocab_keywords):
        return (
            "找出文本中一個有不同字義或較抽象的詞彙，設計「詞義理解」的單選題。\n"
            "陷阱設計：(A)上下文正確解釋 (B)字面混淆 (C)常理干擾 (D)另一個錯誤選項"
        )
    cause_keywords = ['因果', '原因', '結果', '觀點', '理由', '找原因', '找結果']
    if any(k in strategy for k in cause_keywords):
        return (
            "設計「因果關係」或「觀點立場」的單選題。\n"
            "陷阱設計：(A)正確因果 (B)張冠李戴 (C)因果倒置 (D)另一個錯誤選項"
        )
    summary_keywords = ['重點句', '大意', '歸納', '刪除細節', '統整', '主旨', '段落整理']
    if any(k in strategy for k in summary_keywords) or process == '詮釋整合':
        return (
            "設計「段落大意」或「全文主旨」的單選題。\n"
            "陷阱設計：(A)正確核心觀念 (B)以偏概全（細節） (C)過度擴充 (D)另一個錯誤選項"
        )
    predict_keywords = ['預測', '提問', '自我', '推測', '想像']
    if any(k in strategy for k in predict_keywords) or process == '比較評估':
        return (
            "設計「預測發展」或「評估合理性」的單選題。\n"
            "陷阱設計：(A)合理推論 (B)憑空想像 (C)方向相反 (D)另一個錯誤選項"
        )
    if process == '提取訊息':
        return (
            "設計「找出正確敘述」的單選題。\n"
            "陷阱設計：(A)正確陳述 (B)關鍵細節錯誤 (C)人物/時間/地點有誤 (D)另一個錯誤選項"
        )
    return f"設計一道符合「{process}」認知歷程的單選題，四個選項需有明顯教學鑑別力。"


def build_prompt(orig_title, orig_q, transfer_text, strategy, process, grade):
    instruction = get_strategy_instruction(strategy, process)
    text_preview = str(transfer_text)[:800] + ("..." if len(str(transfer_text)) > 800 else "")

    orig_q_display = (
        str(orig_q)
        if orig_q and "(針對" not in str(orig_q) and "ＯＯＯ" not in str(orig_q)
        else "(無原題參考，請依補充文本自行設計)"
    )

    return f"""你是台灣國小{grade}的國語科命題老師。

【原始案例】: {orig_title}
【原始題目（請嚴格對齊此題的難度、提問風格與選項長度）】:
{orig_q_display}

【補充文本（請根據此文出題）】:
{text_preview}

【出題策略】: {strategy}
【出題指示】: {instruction}

注意事項：
1. 題幹風格需與原始題目相似
2. 選項長度和複雜度需和原題相當（簡短對應簡短，詳細對應詳細）
3. 陷阱選項需符合「{strategy}」策略的誘答邏輯

請直接輸出（不需解釋）：
題目：[題幹]
(A) [選項A]
(B) [選項B]
(C) [選項C]
(D) [選項D]
答案：(X)"""


def parse_response(text: str) -> str:
    """Parse Gemini response and format it consistently."""
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    return '\n'.join(lines)


# ===================================
# 主程式
# ===================================
print(f"[OK] Loading Excel: {EXCEL_FILE}")
df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, dtype=str).fillna('')
print(f"[OK] Loaded {len(df)} rows from '{SHEET_NAME}'")

wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

headers = [cell.value for cell in ws[1]]
try:
    q_col = headers.index('遷移題目') + 1
    print(f"[OK] Column '遷移題目' at column {q_col}")
except ValueError:
    print("[ERROR] Column '遷移題目' not found!")
    print("Available columns:", headers)
    exit(1)

total = len(df)
success = 0
skipped = 0
errors = 0

print(f"\n[START] Generating questions for {total} rows...")
print(f"        Saving every {SAVE_EVERY} rows. Starting from row {START_FROM_ROW + 1}.\n")

for idx, row in df.iterrows():
    if idx < START_FROM_ROW:
        skipped += 1
        continue

    transfer_text = str(row.get('遷移文本內容', ''))
    if not transfer_text or transfer_text.startswith('無可用'):
        print(f"  [SKIP] Row {idx + 1}: No transfer text")
        skipped += 1
        continue

    orig_title = str(row.get('原始案例標題', ''))
    orig_q     = str(row.get('原始題目', ''))
    strategy   = str(row.get('教學策略', ''))
    process    = str(row.get('認知歷程', ''))
    grade      = str(row.get('年級', ''))

    prompt = build_prompt(orig_title, orig_q, transfer_text, strategy, process, grade)

    for attempt in range(MAX_RETRIES):
        try:
            response = client.models.generate_content(model=MODEL_NAME, contents=prompt)
            formatted = parse_response(response.text)
            ws.cell(row=idx + 2, column=q_col).value = formatted
            success += 1
            print(f"  [OK] Row {idx + 1}/{total} - {orig_title[:20]}...")
            break
        except Exception as e:
            err_str = str(e)
            if '429' in err_str and attempt < MAX_RETRIES - 1:
                wait = (attempt + 1) * 15  # 15s, 30s, 45s
                print(f"  [RATE LIMIT] Row {idx + 1}, waiting {wait}s then retrying...")
                time.sleep(wait)
            else:
                print(f"  [ERR] Row {idx + 1}: {err_str[:120]}")
                errors += 1
                break

    # Save periodically
    if (idx + 1) % SAVE_EVERY == 0:
        wb.save(EXCEL_FILE)
        print(f"\n  [SAVE] Progress saved at row {idx + 1}.\n")

    time.sleep(SLEEP_SECONDS)

# Final save
wb.save(EXCEL_FILE)

print(f"\n{'=' * 50}")
print(f"[DONE] Complete!")
print(f"  Success : {success}")
print(f"  Skipped : {skipped}")
print(f"  Errors  : {errors}")
print(f"{'=' * 50}")
print(f"\nPlease sync '學習遷移題目' sheet to Google Sheets.")
