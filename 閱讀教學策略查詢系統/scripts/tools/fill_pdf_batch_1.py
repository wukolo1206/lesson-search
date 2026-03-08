import openpyxl
import os
import sys

# 設定路徑
EXCEL_PATH = r'd:\test ch\閱讀教學策略查詢系統\學習扶助國語教材\教材資料庫完整版.xlsx'

# 硬編碼寫死 AI 提煉出的內容
BATCH_1_DATA = {
    "3-1-6": {
        "核心學習領域": "字詞／字詞常識／（四）運用簡單字詞知識輔助讀寫",
        "教學重點摘要": "引導學生認識名詞和動詞，理解兩者是構句的基本要素，並簡單區分及物與不及物動詞。透過圈、寫等讀寫活動找出名詞與動詞，體會詞性學習能輔助增進閱讀理解。",
        "適用考點類型": "語詞理解與運用\n字義與詞性辨識",
        "備課重點提醒": "教學建議以「支援前線」（拿取物品理解名詞）與「比手畫腳」（表演理解動詞）引起動機。發展活動透過「配對遊戲」加深詞性組合概念（如動詞+名詞的完整語意）。最後透過詩歌〈小船〉與〈紅毛城〉統整所學。"
    },
    "3-1-7": {
        "核心學習領域": "字詞／字詞常識／（四）運用簡單字詞知識輔助讀寫",
        "教學重點摘要": "引導學生在生活經驗中發現形容詞「指稱特定事物」的功能，歸納成「形容事物性質/樣態」與「心理感覺」兩大類。透過短文閱讀比較，感受形容詞在修飾語句時所發揮的效果。",
        "適用考點類型": "語詞理解與運用",
        "備課重點提醒": "準備活動可利用「猜盒子、選玩具」凸顯形容詞的重要性；發展環節加入「奇趣箱摸水果」活動，累積如粗粗的、毛毛的等視覺與觸覺感官詞彙；並引導學生在課文語句加上「很、非常」等程度副詞感受其強調效果。"
    },
    "3-3-1": {
        "核心學習領域": "篇章／閱讀／（一）正確而流暢的朗讀文本，並注意語調變化",
        "教學重點摘要": "帶領學生回顧正確切分語句的方法，理解各種標點符號的停頓與表情作用（如問號、驚嘆號），以及應用聲音輕重快慢的變化。進階引導學生推敲出各段落的感覺以及整體故事的情感變化。",
        "適用考點類型": "閱讀理解\n語氣與情緒推論",
        "備課重點提醒": "本單元適合接續在「文本深究」之後教學。引導時可結合過去學過的《句子切一切》、《魔豆貼》等舊經驗。可設計「角色讀」的三人小組活動（扮演小松鼠、啄木鳥、旁白），增添朗讀過程的多樣性與代入感。"
    },
    "3-3-3": {
        "核心學習領域": "篇章／閱讀／（五）運用推測、推論、提問等策略，增進文本理解",
        "教學重點摘要": "引導學生閱讀篇章時主動提出疑問，掌握主要人物和事件間的關聯。重點引導學生根據課文中描述人物的細節，從「語言、表情、動作」三大向度提取訊息，自己推論並詮釋出人物的鮮明個性。",
        "適用考點類型": "詮釋整合\n推論訊息\n人物特質分析",
        "備課重點提醒": "設計「邊聽邊記」的方式，在聆聽《神射手與賣油翁》時用便利貼記錄「何人、何事、如何做」。學生發言銓釋個性時，教師可適時提供輔助詞彙（如：本領高強、驕傲、謙虛）。另建議搭配「倒水體驗」深化賣油翁技術難度的情境感知。"
    },
    "4-1-1": {
        "核心學習領域": "字詞／認識／（二）認識常用語詞1,700個",
        "教學重點摘要": "中年級字詞核心目標在於「自覺不懂詞彙並運用策略推測」。引導學生在文本中遭遇難詞或不合理處時，懂得放慢速度，並利用前後文查找相近/相反詞、句型觀察或「語詞伸展操(擴詞)」等策略來判讀詞義。",
        "適用考點類型": "語詞理解與運用\n推論訊息",
        "備課重點提醒": "可利用《小車站旁的五味屋》為例，針對如「價值非凡」、「陳列」、「繁華褪盡」等書面語進行拆解示範；同時鼓勵學生尋找不同段落中重複出現的線索詞（如「商品」），訓練其運用上下文推論、替換詞語的語感。"
    }
}

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # 找 Header
    header_row_idx = None
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == '教材編號':
            header_row_idx = i
            headers = list(row)
            break
            
    if header_row_idx is None:
        print("找不到教材編號欄位")
        return

    # 定義欄位 index
    col_domain = headers.index('核心學習領域') + 1
    col_summary = headers.index('教學重點摘要') + 1
    col_type = headers.index('適用考點類型') + 1
    col_note = headers.index('備課重點提醒') + 1

    updated_count = 0
    
    for row_idx in range(header_row_idx + 2, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=1).value
        if cell_id and str(cell_id).strip() in BATCH_1_DATA:
            target_id = str(cell_id).strip()
            data = BATCH_1_DATA[target_id]
            
            ws.cell(row=row_idx, column=col_domain).value = data["核心學習領域"]
            ws.cell(row=row_idx, column=col_summary).value = data["教學重點摘要"]
            ws.cell(row=row_idx, column=col_type).value = data["適用考點類型"]
            ws.cell(row=row_idx, column=col_note).value = data["備課重點提醒"]
            
            print(f"成功寫入: {target_id}")
            updated_count += 1

    if updated_count > 0:
        wb.save(EXCEL_PATH)
        print(f"\nBatch 1: 共 {updated_count} 筆資料已成功寫入 `{EXCEL_PATH}`。")
    else:
        print("沒有任何資料被寫入，請檢查編號對應。")

if __name__ == "__main__":
    main()
