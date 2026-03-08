"""
匯出未審查的遷移題目，整理成可貼入 Claude 對話的格式
用法：
  python scripts/tools/export_for_review.py --batch 5    # 匯出 5 題（預設）
  python scripts/tools/export_for_review.py --batch 10   # 匯出 10 題
  python scripts/tools/export_for_review.py --start 50   # 從第 50 列開始找
"""

import openpyxl
import argparse

EXCEL_PATH = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
SHEET_NAME = "學習遷移題目"
OUTPUT_PATH = r"D:\test ch\閱讀教學策略查詢系統\output\review_batch.txt"


def load_original_articles(wb):
    articles = {}
    for sheet_name in ["三年級", "四年級", "五年級", "六年級"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        title_col = next((i for i, h in enumerate(headers) if h and "標題" in str(h)), None)
        text_col = next((i for i, h in enumerate(headers) if h and "全文" in str(h)), None)
        if title_col is None or text_col is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[title_col]
            text = row[text_col]
            if title and text:
                articles[str(title).strip()] = str(text)[:800]
    return articles


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", type=int, default=5, help="一次匯出幾題")
    parser.add_argument("--start", type=int, default=2, help="從哪一列開始搜尋（預設從第2列）")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    print("載入原始文章...")
    original_articles = load_original_articles(wb)
    print(f"  已載入 {len(original_articles)} 篇文章")

    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    collected = []
    row_ids = []

    for row_num in range(args.start, ws.max_row + 1):
        if len(collected) >= args.batch:
            break

        row = [ws.cell(row=row_num, column=i + 1).value for i in range(len(headers))]
        data = dict(zip(headers, row))

        # 跳過已審查
        if data.get("AI_綜合判定"):
            continue

        # 跳過遷移題目為空
        transfer_q = str(data.get("遷移題目") or "").strip()
        if not transfer_q or "ＯＯＯ" in transfer_q:
            continue

        orig_title = str(data.get("原始案例標題") or "").strip()
        original_article = original_articles.get(orig_title, "（原始文章未找到）")
        transfer_text = str(data.get("遷移文本內容") or "").strip()[:3000]

        collected.append({
            "row": row_num,
            "orig_title": orig_title,
            "strategy": str(data.get("教學策略") or ""),
            "cognitive": str(data.get("認知歷程") or ""),
            "original_q": str(data.get("原始題目") or ""),
            "original_article": original_article,
            "transfer_text": transfer_text,
            "transfer_q": transfer_q,
        })
        row_ids.append(row_num)

    if not collected:
        print("沒有找到待審查的題目。")
        return

    # 輸出格式
    lines = []
    lines.append("=" * 60)
    lines.append(f"待審查遷移題目（共 {len(collected)} 題）")
    lines.append(f"列號：{row_ids}")
    lines.append("=" * 60)
    lines.append("")
    lines.append("請依以下格式審查每題，輸出純 JSON 陣列：")
    lines.append("""[
  {
    "row": 列號,
    "答案": "A/B/C/D",
    "答案依據": "引用文章原句",
    "考點同構": 1-5,
    "誘答一致": 1-5,
    "策略適用": 1-5,
    "基礎適配": 1-5,
    "綜合判定": "通過/需微調/建議重作",
    "修改建議": "說明",
    "修正題目": "若需修改則輸出完整題目，否則填空字串"
  },
  ...
]""")
    lines.append("")
    lines.append("審查標準：")
    lines.append("- 考點同構：遷移題是否複製原題解題複雜度與路徑")
    lines.append("- 誘答一致：錯誤選項是否對應相同類別的學生迷思")
    lines.append("- 策略適用：學生是否必須運用該教學策略才能解題")
    lines.append("- 基礎適配：詞彙難度是否適合學扶學生，情境是否有足夠表面變異")
    lines.append("- 認知歷程：提取訊息=答案直接在文章、推論訊息=需串聯兩句以上、詮釋整合=需整體閱讀、比較評估=需個人判斷")
    lines.append("")
    lines.append("-" * 60)

    for item in collected:
        lines.append("")
        lines.append(f"【題目 Row {item['row']}】")
        lines.append(f"原始文章標題：{item['orig_title']}")
        lines.append(f"教學策略：{item['strategy']}")
        lines.append(f"認知歷程：{item['cognitive']}")
        lines.append("")
        lines.append("▌原始文章（節錄）：")
        lines.append(item['original_article'])
        lines.append("")
        lines.append("▌原始題目：")
        lines.append(item['original_q'])
        lines.append("")
        lines.append("▌遷移文本：")
        lines.append(item['transfer_text'])
        lines.append("")
        lines.append("▌待審查遷移題目：")
        lines.append(item['transfer_q'])
        lines.append("")
        lines.append("-" * 60)

    output = "\n".join(lines)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(output)

    print(f"\n已匯出 {len(collected)} 題到：{OUTPUT_PATH}")
    print(f"列號：{row_ids}")
    print("\n請將該檔案內容貼入 Claude 對話進行審查。")
    print("審查完成後，執行 import_review_result.py 寫回 Excel。")


if __name__ == "__main__":
    main()
