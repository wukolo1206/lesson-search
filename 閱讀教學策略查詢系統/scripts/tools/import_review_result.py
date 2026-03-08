"""
將 Claude 審查結果（JSON 陣列）寫回 Excel
用法：
  python scripts/tools/import_review_result.py --file output/result.json
  python scripts/tools/import_review_result.py --json '[{"row":5,...},...]'
"""

import openpyxl
import json
import argparse

EXCEL_PATH = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
SHEET_NAME = "學習遷移題目"

NEW_COLS = [
    "AI_審查題目",
    "AI_答案",
    "AI_答案依據",
    "AI_考點同構",
    "AI_誘答一致",
    "AI_策略適用",
    "AI_基礎適配",
    "AI_綜合判定",
    "AI_修改建議",
]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="JSON 結果檔案路徑")
    parser.add_argument("--json", help="直接傳入 JSON 字串")
    args = parser.parse_args()

    if args.file:
        with open(args.file, "r", encoding="utf-8") as f:
            results = json.load(f)
    elif args.json:
        results = json.loads(args.json)
    else:
        print("請用 --file 或 --json 提供審查結果")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in ws[1]]
    for col_name in NEW_COLS:
        if col_name not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col_name)
            headers.append(col_name)

    col_index = {h: i + 1 for i, h in enumerate(headers)}

    success = 0
    for r in results:
        row_num = r.get("row")
        if not row_num:
            print(f"缺少 row 欄位，跳過：{r}")
            continue

        # 取遷移題目原文（若通過用原文，否則用修正版）
        transfer_q_col = col_index.get("遷移題目")
        original_transfer_q = ws.cell(row=row_num, column=transfer_q_col).value or ""
        reviewed_q = r.get("修正題目", "") or original_transfer_q

        ws.cell(row=row_num, column=col_index["AI_審查題目"]).value = reviewed_q
        ws.cell(row=row_num, column=col_index["AI_答案"]).value = r.get("答案", "")
        ws.cell(row=row_num, column=col_index["AI_答案依據"]).value = r.get("答案依據", "")
        ws.cell(row=row_num, column=col_index["AI_考點同構"]).value = r.get("考點同構", "")
        ws.cell(row=row_num, column=col_index["AI_誘答一致"]).value = r.get("誘答一致", "")
        ws.cell(row=row_num, column=col_index["AI_策略適用"]).value = r.get("策略適用", "")
        ws.cell(row=row_num, column=col_index["AI_基礎適配"]).value = r.get("基礎適配", "")
        ws.cell(row=row_num, column=col_index["AI_綜合判定"]).value = r.get("綜合判定", "")
        ws.cell(row=row_num, column=col_index["AI_修改建議"]).value = r.get("修改建議", "")

        print(f"Row {row_num}：{r.get('綜合判定', '?')} 已寫入")
        success += 1

    wb.save(EXCEL_PATH)
    print(f"\n完成：{success} 題已寫回 Excel。")


if __name__ == "__main__":
    main()
