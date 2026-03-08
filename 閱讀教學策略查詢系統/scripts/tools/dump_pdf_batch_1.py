import openpyxl
import os
import pdfplumber
import sys

# 設定路徑
EXCEL_PATH = r'd:\test ch\閱讀教學策略查詢系統\學習扶助國語教材\教材資料庫完整版.xlsx'
PDF_DIR = r'd:\test ch\閱讀教學策略查詢系統\學習扶助國語教材'
OUTPUT_FILE = r'd:\test ch\閱讀教學策略查詢系統\output\pdf_batch_1.txt'

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # 找 Header
    header_row_idx = None
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == '教材編號':
            header_row_idx = i
            headers = row
            break
            
    if header_row_idx is None:
        print("找不到教材編號欄位")
        return

    # 找出缺資料的編號 (檢查第5欄 '核心學習領域' 或 第6欄 '教學重點摘要' 是否為空)
    missing_ids = []
    for idx, row in enumerate(ws.iter_rows(min_row=header_row_idx+2, values_only=True)):
        if row[0]:
            # index 4 是核心學習領域, 5 是教學重點摘要
            if (len(row) <= 4 or not row[4]) or (len(row) <= 5 or not row[5]):
                missing_ids.append(str(row[0]).strip())

    print(f"共有 {len(missing_ids)} 筆資料缺漏。")
    
    # 取前 5 筆作為 Batch 1
    batch_1_ids = missing_ids[:5]
    print(f"Batch 1 處理名單: {batch_1_ids}")

    all_pdfs = [f for f in os.listdir(PDF_DIR) if f.endswith('.pdf')]
    
    batch_content = []
    
    for target_id in batch_1_ids:
        # 找尋對應的 PDF 檔案 (以 target_id 開頭)
        target_pdf = None
        for f in all_pdfs:
            if f.startswith(f"{target_id}-") or f.startswith(f"{target_id} "):
                target_pdf = f
                break
            # 容錯處理：例如 4-3-1誰是順風耳
            if target_id in f:
                target_pdf = f
                break

        if not target_pdf:
            print(f"找不到 {target_id} 的 PDF 檔案。")
            continue
            
        pdf_path = os.path.join(PDF_DIR, target_pdf)
        print(f"正在解析: {target_pdf}")
        
        text = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            print(f"解析 {target_pdf} 失敗: {e}")
            continue
            
        batch_content.append(f"========== 教材編號：{target_id} ==========\n")
        batch_content.append(f"檔案名稱：{target_pdf}\n\n")
        batch_content.append(text)
        batch_content.append("\n\n")

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.writelines(batch_content)
        
    print(f"\nBatch 1 的 PDF 內容已匯出至 {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
