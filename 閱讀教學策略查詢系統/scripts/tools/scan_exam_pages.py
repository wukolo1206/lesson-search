"""
scan_exam_pages.py
掃描考古題 PDF，找出每篇文章標題對應的頁碼
輸出：output/exam_page_map.json  格式: {"114_G3_小光與雜貨店": 3, ...}
"""
import sys, io, os, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pdfplumber
import openpyxl

BASE = "d:/test ch/閱讀教學策略查詢系統"
EXAM_DIR = os.path.join(BASE, "學習扶助考古題")
XLSX = os.path.join(BASE, "學習扶助閱讀測驗試題分析.xlsx")
OUTPUT = os.path.join(BASE, "output/exam_page_map.json")

# G格式 → 資料夾名稱（模糊比對用）
GRADE_FOLDER = {
    'G3': '三年級',
    'G4': '四年級',
    'G5': '五年級',
    'G6': '六年級',
}

# 年度數字 → 開頭字串比對
def year_matches(fname, year):
    """判斷 PDF 檔名是否屬於指定年度"""
    fname = os.path.basename(fname)
    yr = str(year)
    # 格式：114年、114年...、10805（108年）、109.05、109年
    if fname.startswith(yr + '年') or fname.startswith(yr + '.') or fname.startswith(yr[1:] + '0' + yr[-1]):
        return True
    # 108 年特殊：10805
    if yr == '108' and fname.startswith('10805'):
        return True
    return False

def get_pdf_path(year, grade_g):
    """找出對應的 PDF 檔案路徑"""
    folder_keyword = GRADE_FOLDER.get(grade_g, '')
    if not folder_keyword:
        return None
    # 找包含年級關鍵字的子資料夾
    for folder in os.listdir(EXAM_DIR):
        if folder_keyword in folder:
            folder_path = os.path.join(EXAM_DIR, folder)
            for fname in os.listdir(folder_path):
                if fname.endswith('.pdf') and year_matches(fname, year):
                    return os.path.join(folder_path, fname)
    return None

def search_title_in_pdf(pdf_path, title):
    """在 PDF 各頁搜尋標題關鍵字，回傳頁碼（1-based），找不到回傳 None"""
    # 用標題前 2~4 個字元作為搜尋關鍵字（避免標點等雜訊）
    keywords = []
    clean = re.sub(r'[（）「」『』【】、。，？！《》\s]', '', title)
    if len(clean) >= 4:
        keywords.append(clean[:4])
    if len(clean) >= 3:
        keywords.append(clean[:3])
    if len(clean) >= 2:
        keywords.append(clean[:2])

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ''
                for kw in keywords:
                    if kw in text:
                        return page_num
    except Exception as e:
        print(f'  ERROR reading {os.path.basename(pdf_path)}: {e}')
    return None

def main():
    # 讀取 xlsx，取得所有 (年度, 年級G格式, 文本標題)
    wb = openpyxl.load_workbook(XLSX)
    articles = []  # (year, grade_g, title)

    grade_to_g = {'三年級': 'G3', '四年級': 'G4', '五年級': 'G5', '六年級': 'G6'}

    for sheet_name, g in [('三年級','G3'),('四年級','G4'),('五年級','G5'),('六年級','G6')]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        if '年度' not in headers or '文本標題' not in headers:
            continue
        year_idx = headers.index('年度')
        title_idx = headers.index('文本標題')
        seen = set()
        for row in ws.iter_rows(min_row=2):
            year = row[year_idx].value
            title = row[title_idx].value
            if year and title:
                key = (int(year), g, str(title).strip())
                if key not in seen:
                    seen.add(key)
                    articles.append(key)

    print(f'共 {len(articles)} 篇文章需要掃描\n')

    result = {}
    not_found = []

    for year, grade_g, title in sorted(articles):
        pdf_path = get_pdf_path(year, grade_g)
        if not pdf_path:
            print(f'  [找不到PDF] {year} {grade_g} {title}')
            not_found.append(f'{year}_{grade_g}_{title}')
            continue

        page = search_title_in_pdf(pdf_path, title)
        key = f'{year}_{grade_g}_{title}'
        if page:
            result[key] = page
            print(f'  OK  {year} {grade_g} {title:16s} → 第 {page} 頁  ({os.path.basename(pdf_path)})')
        else:
            result[key] = None
            not_found.append(key)
            print(f'  ??? {year} {grade_g} {title:16s} → 找不到  ({os.path.basename(pdf_path)})')

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    found = sum(1 for v in result.values() if v)
    print(f'\n=== 完成 ===')
    print(f'找到: {found}/{len(result)} 篇')
    print(f'結果儲存: {OUTPUT}')
    if not_found:
        print(f'\n未找到:')
        for x in not_found:
            print(f'  {x}')

if __name__ == '__main__':
    main()
