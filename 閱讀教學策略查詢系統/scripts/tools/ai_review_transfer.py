"""
AI 自動審查遷移題目（Claude Sonnet）
- 讀取「學習遷移題目」工作表中的現有遷移題目
- 同時讀取原始文章全文（從三~六年級工作表）
- 呼叫 Claude Sonnet API 審查並在需要時輸出修正版
- 將結果寫回 Excel 新增欄位
用法：
  python scripts/tools/ai_review_transfer.py --rows 3        # 跑前 3 題（測試）
  python scripts/tools/ai_review_transfer.py --rows all      # 跑全部
  python scripts/tools/ai_review_transfer.py --rows all --force  # 強制重跑已有結果的題目
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

import anthropic
import openpyxl
import json
import argparse
import time
import datetime
import os

EXCEL_PATH = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
SHEET_NAME = "學習遷移題目"
API_KEY = os.environ.get("ANTHROPIC_API_KEY", "YOUR_API_KEY_HERE")

NEW_COLS = [
    "AI_審查題目",    # 若需修改，放修正版；通過則放原題
    "AI_答案",
    "AI_答案依據",    # 引用原文的依據句
    "AI_考點同構",
    "AI_誘答一致",
    "AI_策略適用",
    "AI_基礎適配",
    "AI_綜合判定",
    "AI_修改建議",
]

PROMPT_TEMPLATE = """你是一位國小閱讀理解測驗的出題與評量專家，專門服務「學習扶助」計畫。

## 你的任務
審查「遷移題目」是否達到與「原始題目」相同的測驗品質，並給出評分與建議。

---

## 原始資料

### 原始文章（部分）
{original_article}

### 原始題目
- 教學策略：{strategy}
- 認知歷程：{cognitive}
- 題目：{original_q}

---

## 遷移資料

### 遷移文本（學生閱讀的新文章）
{transfer_text}

### 待審查的遷移題目
{transfer_q}

---

## 審查標準

1. **考點同構**：遷移題是否複製了原題的解題複雜度與路徑？
2. **誘答一致**：錯誤選項是否對應原題相同類別的學生迷思（如：望文生義、過度推論、局部事實）？
3. **策略適用**：學生是否必須運用該教學策略（{strategy}）才能解題？
4. **基礎適配**：詞彙難度是否適合學扶學生？情境是否有足夠表面變異（非只換名字）？

### 認知歷程判斷原則
- **提取訊息**：答案直接寫在文章一句話裡，可以圈出來
- **推論訊息**：需串聯兩句以上，或理解文字背後的隱含意思
- **詮釋整合**：需整體閱讀，掌握主旨、段落關係或作者意圖
- **比較評估**：需帶入個人判斷，或評估多方說法對錯

---

## 輸出格式（純 JSON，禁止加 markdown 或其他文字）
{{
  "答案": "A",
  "答案依據": "引用文章中支持答案的原句",
  "考點同構": 4,
  "誘答一致": 3,
  "策略適用": 5,
  "基礎適配": 4,
  "綜合判定": "需微調",
  "修改建議": "具體說明哪個選項有問題及如何修改",
  "修正題目": "若綜合判定為需微調或建議重作，在此輸出完整修正後題目（含選項）；若通過則填空字串"
}}

規則：
- 綜合判定只填「通過」「需微調」「建議重作」其中之一
- 答案填 A/B/C/D 或 1/2/3/4（依題目格式）
- 各評分為 1-5 的整數
- 修正題目若不需修改則填 ""
"""


def load_original_articles(wb):
    """從三~六年級工作表載入所有原始文章，以標題為 key"""
    articles = {}
    for sheet_name in ["三年級", "四年級", "五年級", "六年級"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if "文本標題" not in headers and "案例標題" not in headers:
            continue
        title_col = headers.index("文本標題") if "文本標題" in headers else headers.index("案例標題")
        text_col = next((i for i, h in enumerate(headers) if h and "全文" in str(h)), None)
        if text_col is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[title_col]
            text = row[text_col]
            if title and text:
                articles[str(title).strip()] = str(text)[:800]
    return articles


def call_claude(client, row_data, original_articles):
    transfer_text = (row_data["遷移文本內容"] or "").replace("1", "").replace("2", "").strip()
    transfer_q = (row_data["遷移題目"] or "").strip()

    # 取得原始文章（找不到則用空字串）
    orig_title = str(row_data.get("原始案例標題") or "").strip()
    original_article = original_articles.get(orig_title, "（原始文章未找到）")

    prompt = PROMPT_TEMPLATE.format(
        original_article=original_article,
        strategy=row_data["教學策略"] or "",
        cognitive=row_data["認知歷程"] or "",
        original_q=row_data["原始題目"] or "",
        transfer_text=transfer_text[:3000],
        transfer_q=transfer_q,
    )

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2048,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    # 若 JSON 被截斷，嘗試自動補上結尾
    if not raw.endswith("}"):
        raw = raw[:raw.rfind('"修改建議"')]
        raw = raw.rstrip(",\n ") + ', "修改建議": "（輸出截斷，請人工確認）"}' if '"修改建議"' not in raw else raw + '"}'
    return json.loads(raw)


def save_md_report(results, start_row, stopped_at, reason):
    """審查結束後儲存 MD 報告"""
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    date_tag = datetime.datetime.now().strftime("%Y%m%d")
    md_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "docs", "reports", f"AI審題進度_{date_tag}.md"
    )
    os.makedirs(os.path.dirname(md_path), exist_ok=True)

    passed = [r for r in results if r["judge"] == "通過"]
    adjust = [r for r in results if "微調" in str(r["judge"])]
    redo = [r for r in results if "重作" in str(r["judge"])]
    errors = [r for r in results if r["judge"] == "錯誤"]

    lines = [
        f"# AI 審題進度報告",
        f"",
        f"- **執行時間**：{now}",
        f"- **審查範圍**：Row {start_row} 起",
        f"- **停止原因**：{reason}",
        f"- **最後審查到**：Row {stopped_at}",
        f"",
        f"## 本次結果統計",
        f"",
        f"| 判定 | 題數 |",
        f"|------|------|",
        f"| 通過 | {len(passed)} |",
        f"| 需微調 | {len(adjust)} |",
        f"| 建議重作 | {len(redo)} |",
        f"| API 錯誤 | {len(errors)} |",
        f"| **合計** | **{len(results)}** |",
        f"",
    ]

    if adjust:
        lines += [f"## 需微調（{len(adjust)} 題）", f""]
        for r in adjust:
            lines.append(f"- **Row {r['row']}**（{r['title']} / {r['cognitive']}）")
            lines.append(f"  - 考點同構={r['scores'][0]} 誘答一致={r['scores'][1]} 策略適用={r['scores'][2]} 基礎適配={r['scores'][3]}")
            lines.append(f"  - 建議：{r['suggestion'][:80]}")
        lines.append("")

    if redo:
        lines += [f"## 建議重作（{len(redo)} 題）", f""]
        for r in redo:
            lines.append(f"- **Row {r['row']}**（{r['title']} / {r['cognitive']}）")
            lines.append(f"  - 建議：{r['suggestion'][:80]}")
        lines.append("")

    if errors:
        lines += [f"## API 錯誤（{len(errors)} 題）", f""]
        for r in errors:
            lines.append(f"- Row {r['row']}：{r['suggestion'][:100]}")
        lines.append("")

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"\n[MD 報告已儲存] {md_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", default="3")
    parser.add_argument("--start-row", type=int, default=2, help="從哪一 row 開始審查（預設 2）")
    parser.add_argument("--force", action="store_true", help="強制重跑已有結果的題目")
    args = parser.parse_args()

    client = anthropic.Anthropic(api_key=API_KEY)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    print("載入原始文章...")
    original_articles = load_original_articles(wb)
    print(f"  已載入 {len(original_articles)} 篇文章")

    headers = [cell.value for cell in ws[1]]
    for col_name in NEW_COLS:
        if col_name not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col_name)
            headers.append(col_name)

    col_index = {h: i + 1 for i, h in enumerate(headers)}
    max_row = ws.max_row
    start_row = args.start_row
    limit = max_row if args.rows == "all" else start_row + int(args.rows) - 1

    success, failed, skipped = 0, 0, 0
    results = []        # 本次審查紀錄（用於 MD 報告）
    stopped_at = start_row
    stop_reason = "正常完成"

    try:
        for row_num in range(start_row, min(limit + 1, max_row + 1)):
            stopped_at = row_num
            row_data = {headers[i]: ws.cell(row=row_num, column=i + 1).value
                        for i in range(len(headers))}

            # 跳過已處理（除非 --force）
            if not args.force:
                existing = ws.cell(row=row_num, column=col_index["AI_綜合判定"]).value
                if existing:
                    print(f"Row {row_num}: 已有結果（{existing}），跳過")
                    skipped += 1
                    continue

            # 跳過遷移題目為空或佔位符的列
            transfer_q = str(row_data.get("遷移題目") or "")
            if not transfer_q or "ＯＯＯ" in transfer_q:
                print(f"Row {row_num}: 遷移題目為空，跳過")
                skipped += 1
                continue

            label = f"{row_data.get('原始案例標題', '')} / {row_data.get('認知歷程', '')}"
            print(f"Row {row_num}: {label} ...", end=" ", flush=True)
            try:
                result = call_claude(client, row_data, original_articles)

                # 若通過放原題，否則放修正版
                reviewed_q = result.get("修正題目", "") or transfer_q

                ws.cell(row=row_num, column=col_index["AI_審查題目"]).value = reviewed_q
                ws.cell(row=row_num, column=col_index["AI_答案"]).value = result.get("答案", "")
                ws.cell(row=row_num, column=col_index["AI_答案依據"]).value = result.get("答案依據", "")
                ws.cell(row=row_num, column=col_index["AI_考點同構"]).value = result.get("考點同構", "")
                ws.cell(row=row_num, column=col_index["AI_誘答一致"]).value = result.get("誘答一致", "")
                ws.cell(row=row_num, column=col_index["AI_策略適用"]).value = result.get("策略適用", "")
                ws.cell(row=row_num, column=col_index["AI_基礎適配"]).value = result.get("基礎適配", "")
                ws.cell(row=row_num, column=col_index["AI_綜合判定"]).value = result.get("綜合判定", "")
                ws.cell(row=row_num, column=col_index["AI_修改建議"]).value = result.get("修改建議", "")

                judge = result.get("綜合判定", "?")
                print(f"[{judge}]")
                success += 1
                results.append({
                    "row": row_num,
                    "title": row_data.get("原始案例標題", ""),
                    "cognitive": row_data.get("認知歷程", ""),
                    "judge": judge,
                    "scores": [
                        result.get("考點同構", ""),
                        result.get("誘答一致", ""),
                        result.get("策略適用", ""),
                        result.get("基礎適配", ""),
                    ],
                    "suggestion": result.get("修改建議", ""),
                })

            except (anthropic.RateLimitError, anthropic.APIStatusError) as e:
                err_msg = str(e)
                is_credit = "credit balance" in err_msg or "too low" in err_msg
                label = "餘額不足" if is_credit else "Quota 耗盡/API 錯誤"
                print(f"[{label}] {err_msg[:120]}")
                stop_reason = f"{label}（Row {row_num}）"
                wb.save(EXCEL_PATH)
                save_md_report(results, start_row, row_num, stop_reason)
                print(f"\n已儲存至 Row {row_num - 1}，下次從 --start-row {row_num} 繼續。")
                return

            except Exception as e:
                print(f"[錯誤] {e}")
                failed += 1
                results.append({
                    "row": row_num,
                    "title": row_data.get("原始案例標題", ""),
                    "cognitive": row_data.get("認知歷程", ""),
                    "judge": "錯誤",
                    "scores": ["", "", "", ""],
                    "suggestion": str(e),
                })

            time.sleep(0.3)

    except KeyboardInterrupt:
        stop_reason = f"使用者中斷（Row {stopped_at}）"

    wb.save(EXCEL_PATH)
    save_md_report(results, start_row, stopped_at, stop_reason)
    print(f"\n完成：成功 {success}｜失敗 {failed}｜跳過 {skipped}，結果已寫回 Excel。")


if __name__ == "__main__":
    main()
