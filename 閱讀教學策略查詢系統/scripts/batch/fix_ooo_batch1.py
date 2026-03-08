import openpyxl
import json

# 親自撰寫前五題的完善題目、選項與答案
QA_BATCH_1 = {
    192: {
        "q": "(針對 7. 識途老馬) 從文章中可以推論出什麼？\n(A) 孤竹國國王對齊桓公很友善\n(B) 老馬的記憶力很好，能循著舊路找到方向\n(C) 管仲是靠觀星象帶領軍隊走出沙漠的\n(D) 齊國軍隊最後在沙漠中全軍覆沒",
        "ans": "解答：(B)"
    },
    195: {
        "q": "(針對 8. 朝三暮四的猴子) 從文章中可以推論出什麼？\n(A) 猴子們其實知道栗子總數沒有改變\n(B) 狙公因為討厭猴子所以想餓死牠們\n(C) 猴子只看眼前數量的變化而被表面假象蒙騙\n(D) 栗子在早上吃比較容易消化",
        "ans": "解答：(C)"
    },
    198: {
        "q": "(針對 1. 逛夜市) 從文章中可以推論出什麼？\n(A) 作者其實一點都不喜歡去夜市\n(B) 夜市通常只有在早上營業\n(C) 夜市是個充滿各式美食小吃及多元遊戲的熱鬧地方\n(D) 夜市裡的遊戲都無法贏得任何獎品",
        "ans": "解答：(C)"
    },
    199: {
        "q": "(針對 1. 逛夜市) 這篇文章主要想告訴我們什麼？\n(A) 介紹臺灣夜市熱鬧的獨特風景與帶給人們的樂趣\n(B) 批評夜市食物不衛生，呼籲大家不要去\n(C) 解釋夜市遊戲機台的運作原理與設計方式\n(D) 抱怨夜市的人潮總是太過擁擠",
        "ans": "解答：(A)"
    },
    201: {
        "q": "(針對 2. 我生病了) 從文章中可以推論出什麼？\n(A) 作者聽從媽媽的話，每天都穿得很保暖\n(B) 媽媽早就提醒過春天天氣變化大，但作者未聽勸而感冒著涼\n(C) 作者生病的原因是因為吃了過期的食物\n(D) 藥袋裡的藥粉全都是甜的，所以生病很舒服",
        "ans": "解答：(B)"
    }
}

def fix_batch():
    wb = openpyxl.load_workbook(r'd:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx')
    ws = wb['學習遷移題目']
    
    headers = [str(cell.value) if cell.value else '' for cell in ws[1]]
    q_idx = headers.index('AI_審查題目')
    ans_idx = headers.index('AI_答案')
    judge_idx = headers.index('AI_綜合判定')
    
    fixed_count = 0
    for row, qa in QA_BATCH_1.items():
        ws.cell(row=row, column=q_idx+1).value = qa['q']
        ws.cell(row=row, column=ans_idx+1).value = qa['ans']
        ws.cell(row=row, column=judge_idx+1).value = '通過 (已透過手動分批更新選項)'
        print(f"Row {row} 完成.")
        fixed_count += 1
        
    wb.save(r'd:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx')
    print(f"批次一完成！共成功寫入 {fixed_count} 筆。")

if __name__ == '__main__':
    fix_batch()
