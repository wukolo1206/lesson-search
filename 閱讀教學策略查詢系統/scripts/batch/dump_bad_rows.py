import openpyxl
import json

wb = openpyxl.load_workbook(r'd:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx', data_only=True)
ws = wb['學習遷移題目']

q_idx = 16 # AI_審查題目
ans_idx = 9 # AI_答案
title_idx = 4 # 遷移文本標題
content_idx = 5 # 遷移文本內容

bad_rows = []
for row in range(409, ws.max_row + 1):
    q = str(ws.cell(row=row, column=q_idx+1).value or '')
    ans = str(ws.cell(row=row, column=ans_idx+1).value or '')
    if 'OOO' in q or 'XXX' in q or 'OOO' in ans or 'XXX' in ans:
        content = str(ws.cell(row=row, column=content_idx+1).value or '')[:400]
        bad_rows.append({
            'row': row,
            'title': ws.cell(row=row, column=title_idx+1).value,
            'q': q,
            'ans': ans,
            'content': content
        })

print(f"Found {len(bad_rows)} bad rows.")
with open('bad_options_new.json', 'w', encoding='utf-8') as f:
    json.dump(bad_rows, f, ensure_ascii=False, indent=2)

for r in bad_rows[:5]:
    print(f"Row: {r['row']}\nTitle: {r['title']}\nQ: {r['q']}\n")
