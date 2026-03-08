import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["學習遷移題目"]
headers = [cell.value for cell in ws[1]]

START, END = 200, 259

for row_num in range(START, END + 1):
    row = {headers[i]: ws.cell(row=row_num, column=i+1).value for i in range(len(headers))}
    print(f"=== Row {row_num} ===")
    print(f"  原始案例標題：{row.get('原始案例標題','')}")
    print(f"  教學策略：{row.get('教學策略','')}")
    print(f"  認知歷程：{row.get('認知歷程','')}")
    print(f"  原始題目：{str(row.get('原始題目',''))[:120]}")
    txt = str(row.get('遷移文本內容','') or '')
    print(f"  遷移文本（前200字）：{txt[:200]}")
    print(f"  遷移題目：{str(row.get('遷移題目',''))[:300]}")
    print(f"  AI_綜合判定（現有）：{row.get('AI_綜合判定','')}")
    print()
