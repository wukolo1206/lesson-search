import openpyxl
import json

wb = openpyxl.load_workbook(r'd:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx', data_only=True)
ws = wb['學習遷移題目']
headers = [str(cell.value) if cell.value else '' for cell in ws[1]]

q_idx = headers.index('AI_生成題目')
alt_q_idx = headers.index('遷移題目')
review_q_idx = headers.index('AI_審查題目')
judge_idx = headers.index('AI_綜合判定')
article_idx = headers.index('遷移文本內容')

bad_options = []
for row_idx in range(2, ws.max_row + 1):
    row = ws[row_idx]
    
    q = ''
    if row[review_q_idx].value:
        q = str(row[review_q_idx].value)
    elif row[q_idx].value:
        q = str(row[q_idx].value)
    elif row[alt_q_idx].value:
        q = str(row[alt_q_idx].value)
        
    if not q: continue
    
    # 檢查是否有 OOO 或 X X X 等字眼
    if 'OOO' in q or 'O O O' in q or 'XXX' in q or 'X X X' in q or 'OXO' in q or 'O X O' in q or 'X O X' in q or 'XOX' in q or 'ＯＯＯ' in q:
        bad_options.append({
            'row': row_idx,
            'q': q,
            'article': str(row[article_idx].value)
        })

print(f'共有 {len(bad_options)} 題含有無效的 OOO 選項。')
with open('bad_options.json', 'w', encoding='utf-8') as f:
    json.dump(bad_options, f, ensure_ascii=False, indent=2)

for item in bad_options[:5]:
    print(f"Row {item['row']}: {item['q'][:50]}...")
