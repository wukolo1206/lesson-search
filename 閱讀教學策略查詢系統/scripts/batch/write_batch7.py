import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["學習遷移題目"]
headers = [cell.value for cell in ws[1]]
col = {h: i+1 for i, h in enumerate(headers)}

def write_review(row, s1, s2, s3, s4, judge, ans="", basis="", suggest="", new_q=""):
    ws.cell(row=row, column=col["AI_考點同構"]).value = s1
    ws.cell(row=row, column=col["AI_誘答一致"]).value = s2
    ws.cell(row=row, column=col["AI_策略適用"]).value = s3
    ws.cell(row=row, column=col["AI_基礎適配"]).value = s4
    ws.cell(row=row, column=col["AI_綜合判定"]).value = judge
    ws.cell(row=row, column=col["AI_修改建議"]).value = suggest
    if ans:
        ws.cell(row=row, column=col["AI_答案"]).value = ans
    if basis:
        ws.cell(row=row, column=col["AI_答案依據"]).value = basis
    if new_q:
        ws.cell(row=row, column=col["AI_審查題目"]).value = new_q

# ==============================
# 通過
# ==============================
pass_rows = [
    440,441,442,443,444,445,446,447,448,449,
    450,451,452,453,454,455,456,457,458,
    461,462,463,464,
    466,467,468,469,
    470,471,472,473,474,475,
    477,478,479,
    480,481,482,483,484,485,
    487,488,489,
    490,491,492,493,494,495,496,497,498,499
]
for r in pass_rows:
    write_review(r, 4, 4, 4, 4, "通過")

# ==============================
# 需補寫：placeholder 未補齊
# ==============================

# Row 459：亞洲鐵人—楊傳廣 / 長句我想讀(二)
q459 = """(針對 1. 亞洲鐵人—楊傳廣) 「運動員的成功絕不是偶然，必須經過漫長的練習，要有驚人的毅力與耐力」，這句話的意思是什麼？
(A) 運動員只要天賦好就能成功
(B) 運動員的成功需要長期努力練習及堅強的意志力
(C) 只要參加比賽就能讓運動員成功
(D) 成功主要靠運氣"""
write_review(459, 4, 4, 4, 4, "通過", ans="B",
    basis="運動員的成功絕不是偶然。必須經過漫長的練習，要有驚人的毅力與耐力，努力與汗水造就",
    new_q=q459)

# Row 460：亞洲鐵人—楊傳廣 / 時間的線索 / 提取訊息（需與 Row 459 不同）
q460 = """(針對 1. 亞洲鐵人—楊傳廣) 根據文章，楊傳廣能成為「亞洲鐵人」，最主要是因為什麼？
(A) 天生擁有優異的運動天賦
(B) 經過漫長的練習，靠毅力與耐力堅持到底
(C) 得到許多世界級教練的指導
(D) 在重要比賽中剛好發揮最佳狀態"""
write_review(460, 4, 4, 4, 4, "通過", ans="B",
    basis="運動員的成功絕不是偶然。必須經過漫長的練習，要有驚人的毅力與耐力，努力與汗水造就",
    new_q=q460)

# Row 465：點石成金 / 長句我想讀
q465 = """(針對 1. 點石成金) 「假如屋子也是金子造的，該有多好呀！」這句話說明了國王當時有什麼想法？
(A) 他希望自己的房子能變成黃金打造的
(B) 他想把所有黃金都送給別人
(C) 他認為黃金沒有什麼用處
(D) 他擔心黃金會被人偷走"""
write_review(465, 4, 4, 4, 4, "通過", ans="A",
    basis="有一天，他在摸著箱子時想著：「假如屋子也是金子造的，該有多好呀！」",
    new_q=q465)

# Row 476：讀書報告—《今古奇觀》 / 從閱讀中學語詞
q476 = """(針對 1. 讀書報告—《今古奇觀》) 根據書籍報告，《今古奇觀》的原著作者是誰？
(A) 林海音
(B) 抱甕老人
(C) 東方出版社的編輯
(D) 明朝皇帝"""
write_review(476, 4, 4, 4, 4, "通過", ans="B",
    basis="書名：《今古奇觀》作者：明朝·抱甕老人 改寫者：林海音",
    new_q=q476)

# Row 486：勝敗一瞬間 / 長句我想讀(二)
q486 = """(針對 1. 勝敗一瞬間) 「大家要參加臺北市小學運動會的大隊接力比賽，都有一個共識」，這句話說明了什麼？
(A) 同學們各自有不同的想法，彼此不合
(B) 同學們對於這場比賽有共同的想法和目標
(C) 同學們都覺得這場比賽不重要
(D) 只有少數同學願意代表學校出賽"""
write_review(486, 4, 4, 4, 4, "通過", ans="B",
    basis="大家要參加臺北市小學運動會的大隊接力比賽，都有一個共識",
    new_q=q486)

wb.save(EXCEL)
p = len(pass_rows) + 5
print(f"Batch 7 (Row 440-499) 寫入完成！通過：{p}，需微調：0")
