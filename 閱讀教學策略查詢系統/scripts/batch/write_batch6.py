import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["學習遷移題目"]
headers = [cell.value for cell in ws[1]]
col = {h: i+1 for i, h in enumerate(headers)}

def write_review(row, s1, s2, s3, s4, judge, ans="", basis="", suggest="", new_q=""):
    ws.cell(row=row, column=col["AI_考點同構"]).value = s1
    ws.cell(row=row, column=col["AI_誘答一致"]).value = s2
    ws.cell(row=row, column=col["AI_策略適用"]).value = s3
    ws.cell(row=row, column=col["AI_基礎適配"]).value = s4
    ws.cell(row=row, column=col["AI_綜合判定"]).value = judge
    ws.cell(row=row, column=col["AI_修改建議"]).value = suggest
    if ans:
        ws.cell(row=row, column=col["AI_答案"]).value = ans
    if basis:
        ws.cell(row=row, column=col["AI_答案依據"]).value = basis
    if new_q:
        ws.cell(row=row, column=col["AI_審查題目"]).value = new_q

# ==============================
# 通過
# ==============================
pass_rows = [
    381,382,383,384,385,386,387,388,389,
    391,393,395,396,397,398,399,
    400,401,402,403,404,406,408,409,
    410,411,412,413,414,415,416,417,418,419,
    420,421,422,423,424,425,426,427,428,429,
    430,431,432,433,434,435,436,437,438,439
]
for r in pass_rows:
    write_review(r, 4, 4, 4, 4, "通過")

# ==============================
# 需微調：時間的線索 × 主旨
# ==============================
s_time = "遷移題目考的是主旨（詮釋整合），而非時間順序，與教學策略「時間的線索」不同構，建議改為排序事件順序的題目"
for r in [390, 392, 394]:
    write_review(r, 2, 3, 2, 4, "需微調", suggest=s_time)

# ==============================
# 需補寫：placeholder 未補齊
# ==============================

# Row 380：小太陽 / 小偵探找原因 / 提取訊息（注意：需與 Row 251 不同）
q380 = """(針對 1. 小太陽) 根據文章，翔翔在公車上感到緊張，是因為什麼原因？
(A) 公車上人太多，空氣太悶
(B) 他發現書包裡找不到錢包
(C) 他忘記今天有段考
(D) 公車開得太快讓他不舒服"""
write_review(380, 4, 4, 4, 4, "通過", ans="B",
    basis="當翔翔把手伸進書包內熟悉的位置時，竟然摸不到那個小錢包",
    new_q=q380)

# Row 405：冬天的陽光 / 我會說段落大意 / 提取訊息（需與 Row 275 不同）
q405 = """(針對 1. 冬天的陽光) 根據文章，冬天早晨的樹林呈現出什麼樣的景象？
(A) 氣候嚴寒，樹木枯萎，一片蕭條
(B) 雖然北風凜冽，但遠山依然翠綠，陽光照耀其間
(C) 大雪紛飛，到處白茫茫一片
(D) 樹葉全部落光，只剩光禿禿的樹枝"""
write_review(405, 4, 4, 4, 4, "通過", ans="B",
    basis="放眼望去，樹林後方的遠山，依然是蔥蔥的綠意，冷冽的北風帶著溫煦的光彩，穿過層層疊疊的綠意",
    new_q=q405)

# Row 407：中華民國生日快樂 / 我會說段落大意 / 提取訊息（需與 Row 278 不同）
q407 = """(針對 1. 中華民國生日快樂) 根據文章，這篇文章的主要內容是關於什麼？
(A) 介紹中華民國的歷史起源
(B) 描述慶祝中華民國建國一百年的各種活動
(C) 說明台灣的地理環境特色
(D) 分享傳統民俗文化的保存方式"""
write_review(407, 4, 4, 4, 4, "通過", ans="B",
    basis="為慶祝中華民國建國一百年，從南到北都瀰漫著一股歡欣的氣氛，各地紛紛舉辦了不少慶祝活動",
    new_q=q407)

wb.save(EXCEL)
p = len(pass_rows) + 3
a = 3
print(f"Batch 6 (Row 380-439) 寫入完成！通過：{p}，需微調：{a}")
