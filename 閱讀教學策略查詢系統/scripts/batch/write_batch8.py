import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["學習遷移題目"]
headers = [cell.value for cell in ws[1]]
col = {h: i+1 for i, h in enumerate(headers)}

def write_review(row, s1, s2, s3, s4, judge, ans="", basis="", suggest="", new_q=""):
    ws.cell(row=row, column=col["AI_考點同構"]).value = s1
    ws.cell(row=row, column=col["AI_誘答一致"]).value = s2
    ws.cell(row=row, column=col["AI_策略適用"]).value = s3
    ws.cell(row=row, column=col["AI_基礎適配"]).value = s4
    ws.cell(row=row, column=col["AI_綜合判定"]).value = judge
    ws.cell(row=row, column=col["AI_修改建議"]).value = suggest
    if ans:
        ws.cell(row=row, column=col["AI_答案"]).value = ans
    if basis:
        ws.cell(row=row, column=col["AI_答案依據"]).value = basis
    if new_q:
        ws.cell(row=row, column=col["AI_審查題目"]).value = new_q

# ==============================
# 通過（500-586 排除已標記行）
# ==============================
skip = {522, 524, 527, 530, 531, 539, 548, 572, 585}
pass_rows = [r for r in range(500, 587) if r not in skip]
for r in pass_rows:
    write_review(r, 4, 4, 4, 4, "通過")

# ==============================
# 需微調：時間的線索 × 主旨
# ==============================
s_time = "遷移題目考的是主旨（詮釋整合），而非時間順序，與教學策略「時間的線索」不同構，建議改為排序事件順序的題目"
for r in [524, 527, 530]:
    write_review(r, 2, 3, 2, 4, "需微調", suggest=s_time)

# ==============================
# 需補寫：placeholder 未補齊
# ==============================

# Row 522：亞洲鐵人 / 段落哪裡轉了彎（與 459/460 不同考點）
q522 = """(針對 1. 亞洲鐵人—楊傳廣) 根據文章，關於楊傳廣成功原因的描述，下列哪個正確？
(A) 他天生就有優異的運動才能，不需要特別訓練
(B) 他透過漫長的練習，靠著毅力與耐力取得傑出成就
(C) 他的成功主要靠著好教練的幫助
(D) 他的成功是運氣使然"""
write_review(522, 4, 4, 4, 4, "通過", ans="B",
    basis="必須經過漫長的練習，要有驚人的毅力與耐力，努力與汗水造就",
    new_q=q522)

# Row 531：點石成金 / 長句我想讀(二)（需與 Row 465 不同長句）
q531 = """(針對 1. 點石成金) 「有一天，他在摸著箱子時想著：『假如屋子也是金子造的，該有多好呀！』話剛說完，奇怪的事情發生了」，這段話說明了什麼？
(A) 國王許下心願之後，奇蹟真的出現了
(B) 國王的房子一直都是用金子蓋的
(C) 國王摸箱子之後感到非常難過
(D) 話說完後什麼事都沒有發生"""
write_review(531, 4, 4, 4, 4, "通過", ans="A",
    basis="有一天，他在摸著箱子時想著：「假如屋子也是金子造的，該有多好呀！」話剛說完",
    new_q=q531)

# Row 539：讀書報告—《今古奇觀》 / 認識說明文 / 圖文轉譯（與 Row 476 不同）
q539 = """(針對 1. 讀書報告—《今古奇觀》) 根據書籍報告，《今古奇觀》是由哪家出版社出版的？
(A) 東方出版社
(B) 林海音出版社
(C) 抱甕老人書坊
(D) 明朝皇家出版社"""
write_review(539, 4, 4, 4, 4, "通過", ans="A",
    basis="出版社：東方出版社",
    new_q=q539)

# Row 548：勝敗一瞬間 / 我會說段落大意（與 Row 486 不同）
q548 = """(針對 1. 勝敗一瞬間) 根據文章，這篇故事主要在描述什麼事情？
(A) 學生練習跑步，準備校內運動會的經過
(B) 代表學校參加大隊接力比賽，體會到勝敗往往在一瞬間的道理
(C) 運動員如何克服緊張情緒贏得冠軍
(D) 老師教導學生如何在比賽中正確傳接棒"""
write_review(548, 4, 4, 4, 4, "通過", ans="B",
    basis="代表學校出賽的大隊接力比賽終於來臨，大家要參加臺北市小學運動會的大隊接力比賽",
    new_q=q548)

# Row 572：文字的魔法 / 時間的線索 / 提取訊息
q572 = """(針對 1. 文字的魔法) 根據文章，「文字沒有生命，是不會說話的，可是只要仙女的魔法棒輕輕一點，它就會活了起來」，文字會活起來的原因是什麼？
(A) 文字本來就有生命，不需要任何幫助
(B) 仙女的魔法棒點了文字之後，文字就有了生命
(C) 努力學習之後，文字自然就會開口說話
(D) 文字只要寫在紙上就可以自己說話"""
write_review(572, 4, 4, 4, 4, "通過", ans="B",
    basis="只要仙女的魔法棒輕輕一點，它就會活了起來",
    new_q=q572)

# Row 585：亞洲鐵人 / 人物放大鏡(一)（與 Row 459/460/522 不同考點）
q585 = """(針對 1. 亞洲鐵人—楊傳廣) 從文章對楊傳廣的描述，可以看出他具有什麼樣的人格特質？
(A) 他隨性自在，不在乎比賽輸贏
(B) 他懶散，靠天賦輕鬆取得好成績
(C) 他努力練習，擁有驚人的毅力和耐力
(D) 他只在乎名次，不重視訓練過程"""
write_review(585, 4, 4, 4, 4, "通過", ans="C",
    basis="必須經過漫長的練習，要有驚人的毅力與耐力，努力與汗水造就",
    new_q=q585)

wb.save(EXCEL)
p = len(pass_rows) + 6
a = 3
print(f"Batch 8 (Row 500-586) 寫入完成！通過：{p}，需微調：{a}")
