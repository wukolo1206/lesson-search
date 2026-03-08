import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["學習遷移題目"]
headers = [cell.value for cell in ws[1]]
col = {h: i+1 for i, h in enumerate(headers)}

def write_review(row, s1, s2, s3, s4, judge, ans="", basis="", suggest="", new_q=""):
    ws.cell(row=row, column=col["AI_考點同構"]).value = s1
    ws.cell(row=row, column=col["AI_誘答一致"]).value = s2
    ws.cell(row=row, column=col["AI_策略適用"]).value = s3
    ws.cell(row=row, column=col["AI_基礎適配"]).value = s4
    ws.cell(row=row, column=col["AI_綜合判定"]).value = judge
    ws.cell(row=row, column=col["AI_修改建議"]).value = suggest
    if ans:
        ws.cell(row=row, column=col["AI_答案"]).value = ans
    if basis:
        ws.cell(row=row, column=col["AI_答案依據"]).value = basis
    if new_q:
        ws.cell(row=row, column=col["AI_審查題目"]).value = new_q

# ==============================
# 通過（大部分）
# ==============================
pass_rows = [
    260,261,262,263,264,265,266,267,268,269,
    270,271,272,273,274,276,277,279,280,
    281,282,283,284,285,286,287,288,289,
    290,291,292,293,295,297,
    300,301,303,304,306,307,308,309,310,
    312,313,314,315,316,317,318,319
]
for r in pass_rows:
    write_review(r, 4, 4, 4, 4, "通過")

# ==============================
# 需微調：長句我想讀 × 比較觀點（294,296,298）
# ==============================
s_long = "策略「長句我想讀」應測驗學生理解長句含義的能力，但遷移題問的是比較兩個觀點，策略不同構，建議改為針對文中長句意思的單選題"
for r in [294, 296, 298]:
    write_review(r, 3, 3, 2, 4, "需微調", suggest=s_long)

# ==============================
# 需微調：學會語詞好閱讀 × 推論（299,302,305）
# ==============================
s_vocab = "策略「學會語詞好閱讀」應測驗學生透過上下文推測詞義，但遷移題問的是一般推論，策略不同構，建議改為詞語意思或用法的選擇題"
for r in [299, 302, 305]:
    write_review(r, 3, 3, 2, 4, "需微調", suggest=s_vocab)

# ==============================
# 需補寫新題（placeholder "通過" 無注記）
# ==============================

# Row 275：冬天的陽光 / 人物放大鏡(一) / 提取訊息
q275 = """(針對 1. 冬天的陽光) 根據文章，作者在冬天的早晨散步時，對眼前的景象有什麼感受？
(A) 覺得冬天景色蕭條，心情低落
(B) 欣賞冬天的美麗景色，感到平靜愉悅
(C) 覺得太冷不舒服，想快點回家
(D) 對眼前景色感到陌生，不知所措"""
write_review(275, 4, 4, 4, 4, "通過", ans="B",
    basis="冷冽的北風帶著溫煦的光彩，穿過層層疊疊的綠意",
    new_q=q275)

# Row 278：中華民國生日快樂 / 人物放大鏡(一) / 提取訊息
q278 = """(針對 1. 中華民國生日快樂) 根據文章，為慶祝建國一百年，各地除了傳統慢跑之外，還舉辦了什麼活動？
(A) 書法展覽
(B) 國慶晚會
(C) 運動會
(D) 花車遊行"""
write_review(278, 4, 4, 4, 4, "通過", ans="B",
    basis="除了傳統慢跑之外，國慶晚會也以令人耳目一新的方式呈現",
    new_q=q278)

# Row 311：夏夜 / 時間的線索 / 提取訊息
q311 = """(針對 1. 夏夜) 根據詩歌，在夏天的夜晚，下列事情發生的順序為何？
甲、蝴蝶和蜜蜂帶著蜜糖回來  乙、羊隊和牛群回家  丙、街燈亮起道晚安  丁、太陽回家
(A) 甲乙丁丙
(B) 丙甲乙丁
(C) 乙甲丁丙
(D) 甲丁乙丙"""
write_review(311, 4, 4, 4, 4, "通過", ans="A",
    basis="蝴蝶和蜜蜂們帶著花朵的蜜糖回來了，羊隊和牛群告別了田野回家了，火紅的太陽也滾著火輪子回家了，當街燈亮起來向村莊道過晚安",
    new_q=q311)

wb.save(EXCEL)
p = len(pass_rows) + 3  # pass + 275/278/311
a = 3 + 3               # 長句 + 語詞
print(f"Batch 4 (Row 260-319) 寫入完成！通過：{p}，需微調：{a}")
