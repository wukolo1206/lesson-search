import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

EXCEL = r"D:\test ch\閱讀教學策略查詢系統\學習扶助閱讀測驗試題分析.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["學習遷移題目"]
headers = [cell.value for cell in ws[1]]
col = {h: i+1 for i, h in enumerate(headers)}

def write_review(row, s1, s2, s3, s4, judge, ans="", basis="", suggest="", new_q=""):
    ws.cell(row=row, column=col["AI_考點同構"]).value = s1
    ws.cell(row=row, column=col["AI_誘答一致"]).value = s2
    ws.cell(row=row, column=col["AI_策略適用"]).value = s3
    ws.cell(row=row, column=col["AI_基礎適配"]).value = s4
    ws.cell(row=row, column=col["AI_綜合判定"]).value = judge
    ws.cell(row=row, column=col["AI_修改建議"]).value = suggest
    if ans:
        ws.cell(row=row, column=col["AI_答案"]).value = ans
    if basis:
        ws.cell(row=row, column=col["AI_答案依據"]).value = basis
    if new_q:
        ws.cell(row=row, column=col["AI_審查題目"]).value = new_q

# ==============================
# 通過
# ==============================
pass_rows = [
    320,321,322,323,324,325,326,327,328,329,
    330,331,332,333,334,335,336,337,338,339,
    340,341,343,345,346,347,348,349,350,351,
    352,353,354,355,357,358,360,361,363,364,
    365,366,367,368,369,370,371,372,373,374,
    375,376,377,378,379
]
for r in pass_rows:
    write_review(r, 4, 4, 4, 4, "通過")

# ==============================
# 需補寫：長句我想讀 placeholder
# ==============================

# Row 342：冬天的陽光 / 長句我想讀 / 提取訊息
q342 = """(針對 1. 冬天的陽光) 「冷冽的北風帶著溫煦的光彩，穿過層層疊疊的綠意」，這句話的意思是什麼？
(A) 冬天的風雖然冷，但陽光溫暖照耀在翠綠的樹林間
(B) 冬天的風很冷，把樹上的葉子都吹光了
(C) 冬天的北風溫暖，讓人感到十分舒適
(D) 冬天的北風擋住了陽光，讓樹林變得昏暗"""
write_review(342, 4, 4, 4, 4, "通過", ans="A",
    basis="冷冽的北風帶著溫煦的光彩，穿過層層疊疊的綠意",
    new_q=q342)

# Row 344：中華民國生日快樂 / 長句我想讀 / 提取訊息
q344 = """(針對 1. 中華民國生日快樂) 「從南到北都瀰漫著一股歡欣的氣氛，各地紛紛舉辦了不少慶祝活動」，這句話說明了什麼？
(A) 全國各地充滿喜悅，並積極舉辦各種慶祝活動
(B) 只有南部地區才有慶祝的氣氛
(C) 各地的慶祝活動規模都很小
(D) 人們普遍不喜歡慶祝節日"""
write_review(344, 4, 4, 4, 4, "通過", ans="A",
    basis="從南到北都瀰漫著一股歡欣的氣氛，各地紛紛舉辦了不少慶祝活動",
    new_q=q344)

# ==============================
# 需微調：學會語詞好閱讀 × 推論
# ==============================
s_vocab = "策略「學會語詞好閱讀」應測驗學生透過上下文推測詞義，但遷移題問的是一般推論，策略不同構，建議改為詞語意思或用法的選擇題"
for r in [356, 359, 362]:
    write_review(r, 3, 3, 2, 4, "需微調", suggest=s_vocab)

wb.save(EXCEL)
p = len(pass_rows) + 2  # pass + 342/344
a = 3
print(f"Batch 5 (Row 320-379) 寫入完成！通過：{p}，需微調：{a}")
